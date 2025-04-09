import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

def export_final_attendance():
    conn = sqlite3.connect('attendance.db')
    cursor = conn.cursor()

    # Step 1: Attendance Records
    cursor.execute("SELECT name, date, time FROM attendance ORDER BY name, date")
    records = cursor.fetchall()

    # Step 2: Summary
    cursor.execute("SELECT name, COUNT(DISTINCT date) FROM attendance GROUP BY name")
    summary_data = dict(cursor.fetchall())
    conn.close()

    if not records:
        print("⚠️ No records found.")
        return

    # Step 3: Main Table
    df = pd.DataFrame(records, columns=["Name", "Date", "Time"])
    df["Days Present"] = df["Name"].map(summary_data)
    file_name = "final_attendance_clean.xlsx"
    df.to_excel(file_name, index=False, sheet_name="Attendance")

    # Step 4: Add hidden summary sheet for chart
    wb = load_workbook(file_name)
    ws_main = wb["Attendance"]
    ws_chart = wb.create_sheet("ChartData")

    ws_chart.sheet_state = 'hidden'  # Hide the sheet

    # Add data to hidden sheet
    ws_chart.cell(row=1, column=1, value="Name")
    ws_chart.cell(row=1, column=2, value="Days Present")
    for i, (name, count) in enumerate(summary_data.items(), start=2):
        ws_chart.cell(row=i, column=1, value=name)
        ws_chart.cell(row=i, column=2, value=count)

    # Step 5: Create chart and add to main sheet
    chart = BarChart()
    chart.title = "Attendance Summary"
    chart.x_axis.title = "Name"
    chart.y_axis.title = "Days Present"
    chart.width = 15
    chart.height = 7

    data = Reference(ws_chart, min_col=2, min_row=1, max_row=1 + len(summary_data))
    cats = Reference(ws_chart, min_col=1, min_row=2, max_row=1 + len(summary_data))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws_main.add_chart(chart, f"E2")

    wb.save(file_name)
    print(f" Exported: '{file_name}' without unnecessary summary table.")

if __name__ == "__main__":
    export_final_attendance()
